"""
Excel Data Analysis AI Agent
Strictly analyzes uploaded Excel/CSV files and answers questions based only on the data
Uses OCRSpace API for image files and Llama via Groq for analysis
"""
import streamlit as st
import pandas as pd
import numpy as np
# import plotly.express as px
# import plotly.graph_objects as go
from typing import TypedDict, Annotated, Sequence
import operator
from langchain_groq import ChatGroq
from langchain_core.messages import BaseMessage, HumanMessage, AIMessage, SystemMessage
from langchain_core.prompts import ChatPromptTemplate
from langgraph.graph import StateGraph, END
from langchain_core.tools import tool
import json
import tempfile
import os
import requests
from PIL import Image
import openpyxl
import time
import concurrent.futures
import re

# Page configuration
st.set_page_config(
    page_title="Excel Data Analysis AI Agent",
    page_icon="📊",
    layout="wide"
)

# Initialize session state
if 'dataframe' not in st.session_state:
    st.session_state.dataframe = None
if 'file_summary' not in st.session_state:
    st.session_state.file_summary = None
if 'extracted_insights' not in st.session_state:
    st.session_state.extracted_insights = None
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []
if 'groq_api_key' not in st.session_state:
    st.session_state.groq_api_key = None
if 'extraction_method' not in st.session_state:
    st.session_state.extraction_method = None
if 'analysis_agent' not in st.session_state:
    st.session_state.analysis_agent = None

# OCR API Configuration
OCR_API_KEY = "K82202332588957"
OCR_API_URL = "https://api.ocr.space/parse/image"

# Analysis Prompt - STRICTLY DATA ONLY
DATA_ANALYSIS_PROMPT = """You are an expert data analyst AI. Your ONLY knowledge source is the uploaded Excel/CSV file data.

## STRICT RULES:
1. **DATA EXCLUSIVITY**: Only use information from the uploaded dataset
2. **NO EXTERNAL KNOWLEDGE**: Do not use any pre-existing knowledge, facts, or general information
3. **REJECT NON-DATA QUESTIONS**: Politely decline questions not about the uploaded data
4. **CITE SPECIFIC DATA**: Reference actual rows, columns, values, and patterns from the data

## Available Dataset Information:
{extracted_info}

## User Question:
{user_question}

## Available Tools:
- get_data_overview: Get overview of all columns and data
- get_column_data: Extract data from specific columns
- search_data_content: Search for specific content in the data
- analyze_data_patterns: Find trends and patterns

## IMPORTANT TOOL USAGE RULES:
1. Only use column names that actually exist in the dataset
2. If unsure about column names, use get_data_overview first
3. Use search_data_content to find specific terms
4. Always verify column names before using them

## Response Format:
- If question is about the data: Provide comprehensive analysis with specific data references
- If data insufficient: "The dataset doesn't contain information to answer this"
- If unrelated: "I can only analyze the uploaded Excel/CSV data"

Provide your data-driven analysis:"""

# Define the state for analysis pipeline
class AgentState(TypedDict):
    messages: Annotated[Sequence[BaseMessage], operator.add]
    dataframe_info: str
    extracted_insights: str
    current_question: str
    analysis_results: dict

# OCR Function
def ocr_space_api(image_file, overlay=False, language='eng'):
    """Extract text from image using OCR.Space API"""
    try:
        payload = {
            'apikey': OCR_API_KEY,
            'language': language,
            'isOverlayRequired': overlay,
            'OCREngine': 2,
            'scale': True,
            'isTable': True,
            'detectOrientation': True
        }
        
        response = requests.post(
            OCR_API_URL,
            files={'image': image_file},
            data=payload,
            timeout=60
        )
        
        result = response.json()
        
        if result.get('IsErroredOnProcessing'):
            return None, f"OCR Error: {result.get('ErrorMessage', 'Unknown error')}"
        
        parsed_results = result.get('ParsedResults', [])
        if not parsed_results:
            return None, "No text found in image"
        
        full_text = ""
        for parsed_result in parsed_results:
            text = parsed_result.get('ParsedText', '')
            if text:
                full_text += text + "\n\n"
        
        return full_text.strip(), None
        
    except Exception as e:
        return None, f"OCR API Error: {str(e)}"

def safe_file_operation(file_path, operation, max_retries=3, delay=1):
    """Safely perform file operations with retries"""
    for attempt in range(max_retries):
        try:
            if operation == 'delete':
                if os.path.exists(file_path):
                    os.unlink(file_path)
                    return True
                return True
        except PermissionError:
            if attempt < max_retries - 1:
                time.sleep(delay)
                continue
            else:
                return False
        except Exception:
            return False
    return False

def extract_excel_data(file_path, sheet_name):
    """Extract data from Excel sheet"""
    try:
        print(f"🔍 Extracting data from sheet: {sheet_name}")
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet_name]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                data.append([str(cell).strip() if cell is not None else "" for cell in row])
        
        wb.close()
        
        if not data:
            print("❌ No data found in sheet")
            return pd.DataFrame()
        
        df = pd.DataFrame(data)
        print(f"📊 Raw data extracted: {df.shape[0]} rows, {df.shape[1]} columns")
        
        # Try to identify header row
        if len(df) > 1:
            header_scores = []
            for i, row in enumerate(data[:5]):
                text_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
                header_scores.append((i, text_count))
            
            header_scores.sort(key=lambda x: x[1], reverse=True)
            if header_scores and header_scores[0][1] >= max(2, len(df.columns) * 0.3):
                best_header_row = header_scores[0][0]
                df.columns = df.iloc[best_header_row]
                df = df.drop(best_header_row).reset_index(drop=True)
                print(f"✅ Header identified at row {best_header_row + 1}")
        
        # Clean the dataframe
        df = df.replace("", np.nan)
        df = df.dropna(how='all').reset_index(drop=True)
        df = df.loc[:, ~df.isna().all()].reset_index(drop=True)
        df = df.fillna("")
        
        print(f"✅ Cleaned data: {df.shape[0]} rows, {df.shape[1]} columns")
        print(f"📋 Columns: {list(df.columns)}")
        
        return df
        
    except Exception as e:
        print(f"❌ Error extracting from sheet '{sheet_name}': {str(e)}")
        st.error(f"Error extracting from sheet '{sheet_name}': {str(e)}")
        return pd.DataFrame()

def extract_all_excel_data(uploaded_file):
    """Extract data from all sheets in Excel file"""
    temp_path = None
    try:
        print("📁 Starting Excel file extraction...")
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            temp_path = tmp_file.name
        
        all_dataframes = {}
        
        # Get sheet names
        try:
            wb = openpyxl.load_workbook(temp_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            print(f"📑 Sheets found: {sheet_names}")
        except Exception as e:
            print(f"❌ Error reading Excel file: {str(e)}")
            st.error(f"Error reading Excel file: {str(e)}")
            return {}
        
        # Process each sheet
        for sheet_name in sheet_names:
            print(f"🔍 Processing sheet: {sheet_name}")
            df = extract_excel_data(temp_path, sheet_name)
            if df is not None and not df.empty:
                all_dataframes[sheet_name] = df
                print(f"✅ Successfully extracted data from {sheet_name}")
            else:
                print(f"⚠️ No data extracted from {sheet_name}")
        
        print(f"📊 Total sheets with data: {len(all_dataframes)}")
        return all_dataframes
        
    except Exception as e:
        print(f"❌ Excel extraction failed: {str(e)}")
        st.error(f"Excel extraction failed: {str(e)}")
        return {}
    finally:
        if temp_path:
            safe_file_operation(temp_path, 'delete')
            print("🧹 Temporary file cleaned up")

def extract_data_insights(api_key: str, df: pd.DataFrame, extraction_method: str) -> str:
    """Extract initial insights from the data"""
    print("🧠 Starting data insights extraction...")
    
    analysis_llm = ChatGroq(
        groq_api_key=api_key,
        model_name="llama-3.3-70b-versatile",
        temperature=0.1,
        max_tokens=3000
    )
    
    # Create comprehensive dataset info
    dataset_info = f"""
**DATASET OVERVIEW:**
- **Source**: {extraction_method}
- **Dimensions**: {df.shape[0]} rows, {df.shape[1]} columns
- **Total Data Points**: {df.shape[0] * df.shape[1]:,}
- **Memory Usage**: {df.memory_usage(deep=True).sum() / 1024**2:.2f} MB

**DATA QUALITY:**
- **Missing Values**: {df.isnull().sum().sum():,} ({(df.isnull().sum().sum() / (df.shape[0] * df.shape[1])) * 100:.1f}%)
- **Duplicate Rows**: {df.duplicated().sum():,}
- **Complete Columns**: {sum(df.isnull().sum() == 0)}/{df.shape[1]}

**COLUMN DETAILS:**
"""
    for col in df.columns:
        null_count = df[col].isnull().sum()
        unique_count = df[col].nunique()
        dataset_info += f"- **{col}**: {df[col].dtype}, {null_count} nulls, {unique_count} unique values\n"
    
    dataset_info += f"""

**SAMPLE DATA (First 10 rows):**
{df.head(10).to_string()}
"""
    
    prompt = f"""Analyze this dataset and extract key structural insights, patterns, and data characteristics:

{dataset_info}

Focus on:
1. Data structure and relationships
2. Key columns and their potential meaning
3. Data quality issues
4. Initial pattern observations
5. Potential analysis opportunities

Provide comprehensive data extraction:"""
    
    print("🤖 Sending request to LLM for insights...")
    messages = [HumanMessage(content=prompt)]
    response = analysis_llm.invoke(messages)
    print("✅ Insights extraction completed")
    
    return response.content

# Define robust analysis tools
@tool
def get_data_overview() -> str:
    """Get overview of all columns and basic statistics"""
    print("🛠️ Tool called: get_data_overview")
    
    if st.session_state.dataframe is None:
        print("❌ No dataframe in session state")
        return "No data loaded."
    
    df = st.session_state.dataframe
    print(f"📊 Processing dataframe with shape: {df.shape}")
    print(f"📋 Available columns: {list(df.columns)}")
    
    overview = {
        "total_columns": len(df.columns),
        "total_rows": len(df),
        "column_names": list(df.columns),
        "data_types": {col: str(df[col].dtype) for col in df.columns},
        "basic_stats": {}
    }
    
    # Add basic statistics for each column
    for col in df.columns:
        col_data = df[col]
        stats = {
            "non_null_count": int(col_data.notna().sum()),
            "null_count": int(col_data.isna().sum()),
            "unique_values": int(col_data.nunique()),
            "sample_values": col_data.dropna().head(3).tolist() if not col_data.empty else []
        }
        
        if pd.api.types.is_numeric_dtype(col_data):
            stats.update({
                "mean": float(col_data.mean()) if not col_data.empty else None,
                "min": float(col_data.min()) if not col_data.empty else None,
                "max": float(col_data.max()) if not col_data.empty else None
            })
        
        overview["basic_stats"][col] = stats
    
    print("✅ get_data_overview completed successfully")
    return json.dumps(overview, default=str)

@tool
def get_column_data(column_name: str) -> str:
    """Extract data from a specific column. Always verify column exists first."""
    print(f"🛠️ Tool called: get_column_data with column_name='{column_name}'")
    
    if st.session_state.dataframe is None:
        print("❌ No dataframe in session state")
        return "No data loaded."
    
    df = st.session_state.dataframe
    print(f"📊 Available columns: {list(df.columns)}")
    
    # Get actual column names for matching
    actual_columns = list(df.columns)
    
    # Try to find the best matching column name
    matched_column = None
    column_name_lower = column_name.lower().strip()
    
    print(f"🔍 Looking for column matching: '{column_name}'")
    
    for actual_col in actual_columns:
        if column_name_lower in actual_col.lower() or actual_col.lower() in column_name_lower:
            matched_column = actual_col
            print(f"✅ Matched '{column_name}' to '{actual_col}'")
            break
    
    # If no close match found, use the first part of the requested name
    if not matched_column and "_" in column_name:
        possible_prefix = column_name.split("_")[0].lower()
        print(f"🔍 Trying prefix matching with: '{possible_prefix}'")
        for actual_col in actual_columns:
            if possible_prefix in actual_col.lower():
                matched_column = actual_col
                print(f"✅ Prefix matched '{column_name}' to '{actual_col}'")
                break
    
    # If still no match, return available columns
    if not matched_column:
        print(f"❌ No match found for '{column_name}'")
        return json.dumps({
            "error": f"Column '{column_name}' not found.",
            "available_columns": actual_columns,
            "suggestions": [col for col in actual_columns if any(word in col.lower() for word in column_name_lower.split())]
        }, default=str)
    
    # Extract data from the matched column
    col_data = df[matched_column]
    print(f"📈 Extracting data from column '{matched_column}', shape: {col_data.shape}")
    
    result = {
        "requested_column": column_name,
        "matched_column": matched_column,
        "total_values": len(col_data),
        "non_null_count": int(col_data.notna().sum()),
        "null_count": int(col_data.isna().sum()),
        "unique_count": int(col_data.nunique()),
        "data_type": str(col_data.dtype),
        "values": col_data.dropna().tolist() if not col_data.empty else []
    }
    
    # Add statistics for numeric columns
    if pd.api.types.is_numeric_dtype(col_data):
        result["statistics"] = {
            "mean": float(col_data.mean()) if not col_data.empty else None,
            "median": float(col_data.median()) if not col_data.empty else None,
            "min": float(col_data.min()) if not col_data.empty else None,
            "max": float(col_data.max()) if not col_data.empty else None,
            "std": float(col_data.std()) if not col_data.empty else None
        }
    else:
        # For non-numeric, show value counts
        value_counts = col_data.value_counts().head(10)
        result["top_values"] = value_counts.to_dict()
    
    print(f"✅ get_column_data completed for '{matched_column}'")
    return json.dumps(result, default=str)

@tool
def search_data_content(search_terms: str) -> str:
    """Search for specific content across all columns"""
    print(f"🛠️ Tool called: search_data_content with search_terms='{search_terms}'")
    
    if st.session_state.dataframe is None:
        print("❌ No dataframe in session state")
        return "No data loaded."
    
    df = st.session_state.dataframe
    search_terms_list = [term.strip().lower() for term in search_terms.split(",")]
    
    print(f"🔍 Searching for terms: {search_terms_list}")
    
    results = {}
    
    for term in search_terms_list:
        term_results = {}
        for col_name in df.columns:
            # Convert column to string for searching
            col_str = df[col_name].astype(str)
            matches = col_str.str.lower().str.contains(term, na=False, regex=False)
            
            if matches.any():
                matching_values = col_str[matches].unique().tolist()
                term_results[col_name] = {
                    "match_count": int(matches.sum()),
                    "sample_matches": matching_values[:5]
                }
                print(f"✅ Found '{term}' in column '{col_name}': {matches.sum()} matches")
        
        if term_results:
            results[term] = term_results
        else:
            results[term] = {"message": f"No matches found for '{term}'"}
            print(f"❌ No matches found for '{term}'")
    
    print("✅ search_data_content completed")
    return json.dumps({"search_results": results}, default=str)

@tool
def analyze_data_patterns() -> str:
    """Analyze overall patterns and trends in the data"""
    print("🛠️ Tool called: analyze_data_patterns")
    
    if st.session_state.dataframe is None:
        print("❌ No dataframe in session state")
        return "No data loaded."
    
    df = st.session_state.dataframe
    patterns = {}
    
    # Basic data patterns
    patterns["data_overview"] = {
        "shape": df.shape,
        "total_cells": df.shape[0] * df.shape[1],
        "missing_cells": df.isnull().sum().sum(),
        "missing_percentage": f"{(df.isnull().sum().sum() / (df.shape[0] * df.shape[1])) * 100:.1f}%"
    }
    
    # Column type analysis
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    text_cols = df.select_dtypes(include=['object']).columns.tolist()
    
    patterns["column_types"] = {
        "numeric_columns": numeric_cols,
        "text_columns": text_cols,
        "total_numeric": len(numeric_cols),
        "total_text": len(text_cols)
    }
    
    # Correlation analysis for numeric columns
    if len(numeric_cols) >= 2:
        try:
            corr_matrix = df[numeric_cols].corr()
            strong_correlations = []
            
            for i in range(len(corr_matrix.columns)):
                for j in range(i+1, len(corr_matrix.columns)):
                    corr_value = corr_matrix.iloc[i, j]
                    if abs(corr_value) > 0.7:
                        strong_correlations.append({
                            "columns": [corr_matrix.columns[i], corr_matrix.columns[j]],
                            "correlation": float(corr_value)
                        })
            
            patterns["correlations"] = strong_correlations
            print(f"📈 Found {len(strong_correlations)} strong correlations")
        except Exception as e:
            patterns["correlations"] = f"Could not calculate correlations: {str(e)}"
            print(f"❌ Correlation calculation failed: {str(e)}")
    
    # Unique value patterns
    patterns["uniqueness"] = {}
    for col in df.columns:
        unique_ratio = df[col].nunique() / len(df[col])
        patterns["uniqueness"][col] = {
            "unique_count": df[col].nunique(),
            "unique_ratio": f"{unique_ratio:.2%}",
            "high_cardinality": unique_ratio > 0.9
        }
    
    print("✅ analyze_data_patterns completed")
    return json.dumps(patterns, default=str)

# Define tools
tools = [
    get_data_overview,
    get_column_data, 
    search_data_content,
    analyze_data_patterns
]

def create_analysis_agent(api_key: str, extracted_insights: str):
    """Create the data analysis agent with robust error handling"""
    print("🤖 Creating analysis agent...")
    
    llm = ChatGroq(
        groq_api_key=api_key,
        model_name="llama-3.3-70b-versatile",
        temperature=0.2,
        max_tokens=4000,
        timeout=30
    )
    
    llm_with_tools = llm.bind_tools(tools)
    print("✅ Tools bound to LLM")
    
    def agent_node(state: AgentState):
        """Main agent node for data analysis with robust error handling"""
        print("🎯 Agent node called")
        messages = state["messages"]
        current_question = state.get("current_question", "")
        
        print(f"❓ User question: '{current_question}'")
        print(f"💬 Messages count: {len(messages)}")
        
        # Strict data-only enforcement
        if st.session_state.dataframe is None:
            print("❌ No dataframe available")
            response_content = "📊 Please upload an Excel or CSV file to begin data analysis."
            return {"messages": [AIMessage(content=response_content)]}
        
        # Enhanced data question detection
        data_keywords = [
            'data', 'analysis', 'analyze', 'trend', 'pattern', 'statistic', 
            'correlation', 'insight', 'column', 'row', 'dataset', 'excel',
            'csv', 'sheet', 'table', 'summary', 'overview', 'values',
            'numbers', 'what does this data', 'tell me about this data',
            'missing', 'duplicates', 'outliers', 'distribution', 'show me',
            'list', 'find', 'search', 'how many', 'what percentage',
            'recommend', 'suggest', 'advise', 'trends', 'patterns',
            'extract', 'get', 'display', 'show'
        ]
        
        question_lower = current_question.lower()
        is_data_question = any(keyword in question_lower for keyword in data_keywords)
        
        print(f"🔍 Data question detected: {is_data_question}")
        
        if not is_data_question:
            response_content = "🔍 I specialize in analyzing uploaded Excel/CSV data. Please ask questions about your dataset, such as trends, patterns, statistics, or insights from the data."
            return {"messages": [AIMessage(content=response_content)]}
        
        # Data analysis path
        system_content = DATA_ANALYSIS_PROMPT.format(
            extracted_info=state["extracted_insights"],
            user_question=current_question
        )
        
        system_msg = SystemMessage(content=system_content)
        
        try:
            print("🤖 Sending request to LLM...")
            response = llm_with_tools.invoke([system_msg] + messages[-3:])
            print("✅ LLM response received")
            
            # Check if response has tool calls
            if hasattr(response, 'tool_calls') and response.tool_calls:
                print(f"🛠️ Tool calls detected: {len(response.tool_calls)}")
                for tool_call in response.tool_calls:
                    print(f"  - Tool: {tool_call['name']}, Args: {tool_call['args']}")
            else:
                print("📝 No tool calls in response - direct answer")
                
            return {"messages": [response]}
        except Exception as e:
            # Fallback response if tool calling fails
            print(f"❌ LLM invocation failed: {str(e)}")
            error_response = f"""I encountered an issue while analyzing your data. Let me provide a direct analysis:

Based on your question about '{current_question}', here's what I can tell you about the dataset:

**Dataset Overview:**
- **Total Rows**: {st.session_state.dataframe.shape[0]:,}
- **Total Columns**: {st.session_state.dataframe.shape[1]}
- **Available Columns**: {list(st.session_state.dataframe.columns)}

**To answer your question specifically:**
Please check if the column names you mentioned exist in the dataset above. You can ask me to:
- "Show all columns" to see available column names
- "Search for [term]" to find specific content
- "Get data from [column name]" to extract specific column data

Would you like me to help you explore the available data in a different way?"""
            
            return {"messages": [AIMessage(content=error_response)]}
    
    # Build the graph
    workflow = StateGraph(AgentState)
    workflow.add_node("agent", agent_node)
    workflow.set_entry_point("agent")
    workflow.add_edge("agent", END)
    
    app = workflow.compile()
    print("✅ Analysis agent created successfully")
    return app

def process_uploaded_file(uploaded_file):
    """Process the uploaded file and extract data"""
    print(f"📁 Processing uploaded file: {uploaded_file.name}, Type: {uploaded_file.type}")
    
    try:
        file_size = len(uploaded_file.getvalue()) / 1024 / 1024
        print(f"📏 File size: {file_size:.2f} MB")
        
        if uploaded_file.type.startswith('image'):
            # Process image files with OCR
            st.info("🖼️ Image file detected - Extracting data using OCR")
            image = Image.open(uploaded_file)
            
            with st.spinner("Extracting data from image..."):
                temp_path = None
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_file:
                        image.save(tmp_file.name)
                        temp_path = tmp_file.name
                    
                    print("🔍 Running OCR...")
                    ocr_text, error = ocr_space_api(open(temp_path, 'rb'))
                    
                    if error:
                        print(f"❌ OCR Error: {error}")
                        st.error(f"OCR Error: {error}")
                        return None
                    
                    # Convert OCR text to simple DataFrame
                    lines = [line for line in ocr_text.split('\n') if line.strip()]
                    if lines:
                        data = []
                        for line in lines:
                            values = [val.strip() for val in line.split() if val.strip()]
                            if values:
                                data.append(values)
                        
                        if data:
                            # Create DataFrame from OCR data
                            max_cols = max(len(row) for row in data)
                            for i, row in enumerate(data):
                                while len(row) < max_cols:
                                    row.append("")
                            
                            df = pd.DataFrame(data)
                            st.session_state.dataframe = df
                            st.session_state.extraction_method = "OCR from Image"
                            print(f"✅ OCR extraction complete! {df.shape[0]} rows, {df.shape[1]} columns")
                            st.success(f"✅ OCR extraction complete! {df.shape[0]} rows, {df.shape[1]} columns")
                            return df
                finally:
                    if temp_path:
                        safe_file_operation(temp_path, 'delete')
            
        else:
            # Process data files
            if uploaded_file.name.endswith('.csv'):
                st.info("📄 CSV file detected")
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file)
                st.session_state.dataframe = df
                st.session_state.extraction_method = "CSV File"
                print(f"✅ CSV loaded! {df.shape[0]:,} rows, {df.shape[1]} columns")
                st.success(f"✅ CSV loaded! {df.shape[0]:,} rows, {df.shape[1]} columns")
                return df
            else:
                st.info("📊 Excel file detected")
                all_dataframes = extract_all_excel_data(uploaded_file)
                
                if not all_dataframes:
                    st.error("❌ No data could be extracted from the Excel file")
                    return None
                
                # Use the largest sheet
                main_sheet = max(all_dataframes.items(), key=lambda x: x[1].shape[0])
                df = main_sheet[1]
                st.session_state.dataframe = df
                st.session_state.extraction_method = f"Excel: {main_sheet[0]}"
                print(f"✅ Excel loaded! {df.shape[0]:,} rows, {df.shape[1]} columns from sheet '{main_sheet[0]}'")
                st.success(f"✅ Excel loaded! {df.shape[0]:,} rows, {df.shape[1]} columns from sheet '{main_sheet[0]}'")
                return df
        
        return None
        
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")
        st.error(f"❌ Error processing file: {str(e)}")
        return None

# Streamlit UI
def main():
    st.title("📊 Excel Data Analysis AI Agent")
    st.markdown("**Upload your Excel/CSV file and ask questions about your data**")
    
    # Sidebar for configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        # API Key input
        api_key = st.text_input(
            "Groq API Key",
            type="password",
            value=st.session_state.groq_api_key or "",
            help="Get your API key from https://console.groq.com"
        )
        
        if api_key:
            st.session_state.groq_api_key = api_key
            st.success("✅ API Key configured")
        else:
            st.warning("⚠️ Please enter your Groq API key")
        
        st.divider()
        
        # File upload
        st.header("📁 Upload Data File")
        uploaded_file = st.file_uploader(
            "Choose Excel or CSV file",
            type=['xlsx', 'xls', 'csv', 'png', 'jpg', 'jpeg'],
            help="Upload your data file for analysis"
        )
        
        if uploaded_file:
            df = process_uploaded_file(uploaded_file)
            
            if df is not None:
                # Show data overview
                st.subheader("📋 Data Overview")
                st.write(f"**File:** {uploaded_file.name}")
                st.write(f"**Rows:** {df.shape[0]:,}")
                st.write(f"**Columns:** {df.shape[1]}")
                st.write(f"**Extraction:** {st.session_state.extraction_method}")
                
                # Show column list
                st.subheader("🗂️ Columns")
                for i, col in enumerate(df.columns):
                    st.write(f"{i+1}. {col}")
                
                # Data insights extraction
                if st.session_state.groq_api_key and st.session_state.extracted_insights is None:
                    st.divider()
                    if st.button("🔍 Analyze Data Patterns", use_container_width=True):
                        with st.spinner("Analyzing data structure and patterns..."):
                            extracted = extract_data_insights(
                                st.session_state.groq_api_key, 
                                df, 
                                st.session_state.extraction_method
                            )
                            st.session_state.extracted_insights = extracted
                            st.session_state.analysis_agent = create_analysis_agent(
                                st.session_state.groq_api_key,
                                extracted
                            )
                            st.success("✅ Data analysis ready!")
                            st.rerun()
    
    # Main content area
    if not st.session_state.groq_api_key:
        st.info("👈 Please enter your Groq API key in the sidebar to get started")
        return
    
    if st.session_state.dataframe is None:
        st.info("📁 Please upload an Excel or CSV file in the sidebar to begin analysis")
        return
    
    # Data preview
    with st.expander("👀 Data Preview", expanded=True):
        df = st.session_state.dataframe
        st.dataframe(df.head(10), use_container_width=True)
        st.write(f"*Showing first 10 of {df.shape[0]:,} rows*")
    
    # Chat interface
    st.divider()
    st.header("💬 Ask Questions About Your Data")
    
    # Display chat history
    for message in st.session_state.chat_history:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])
    
    # Chat input
    if prompt := st.chat_input("Ask about trends, patterns, or insights in your data..."):
        print(f"💬 User input received: '{prompt}'")
        
        # Add user message to chat history
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        
        # Display user message
        with st.chat_message("user"):
            st.markdown(prompt)
        
        # Generate response
        with st.chat_message("assistant"):
            with st.spinner("🔍 Analyzing your data..."):
                try:
                    print("🚀 Starting analysis process...")
                    
                    if st.session_state.extracted_insights is None:
                        response_content = "Please click 'Analyze Data Patterns' in the sidebar first to enable detailed analysis."
                        print("❌ No extracted insights available")
                    else:
                        if st.session_state.analysis_agent is None:
                            print("🤖 Creating analysis agent on demand...")
                            st.session_state.analysis_agent = create_analysis_agent(
                                st.session_state.groq_api_key,
                                st.session_state.extracted_insights
                            )
                        
                        initial_state = {
                            "messages": [HumanMessage(content=prompt)],
                            "dataframe_info": st.session_state.file_summary or "",
                            "extracted_insights": st.session_state.extracted_insights,
                            "current_question": prompt,
                            "analysis_results": {}
                        }
                        
                        print("🎯 Invoking analysis agent...")
                        result = st.session_state.analysis_agent.invoke(initial_state)
                        final_message = result["messages"][-1]
                        response_content = final_message.content
                        print("✅ Analysis completed successfully")
                    
                    # Display response
                    st.markdown(response_content)
                    
                    # Add to chat history
                    st.session_state.chat_history.append({
                        "role": "assistant",
                        "content": response_content
                    })
                    
                except Exception as e:
                    error_msg = f"❌ Analysis error: {str(e)}"
                    print(f"💥 Analysis failed: {str(e)}")
                    st.error(error_msg)
                    
                    # Provide helpful fallback response
                    fallback_response = f"""
I encountered an error while processing your request. Here's what I can tell you about your dataset:

**Dataset Overview:**
- **Total Rows**: {st.session_state.dataframe.shape[0]:,}
- **Total Columns**: {st.session_state.dataframe.shape[1]}
- **Available Columns**: {list(st.session_state.dataframe.columns)}

**Try asking questions like:**
- "Show me all column names"
- "What are the main trends in the data?"
- "Analyze patterns in [specific column]"
- "Search for [specific term] in the data"

Please try rephrasing your question or check if the column names exist in the dataset.
"""
                    st.markdown(fallback_response)
                    st.session_state.chat_history.append({
                        "role": "assistant", 
                        "content": fallback_response
                    })
    
    # Quick analysis suggestions
    if st.session_state.dataframe is not None and len(st.session_state.chat_history) == 0:
        st.divider()
        st.subheader("🚀 Quick Analysis Questions")
        
        col1, col2 = st.columns(2)
        
        quick_questions = [
            ("📈 Data Summary", "Provide a comprehensive summary of this dataset"),
            ("🔍 Show Columns", "Show me all available column names"),
            ("📊 Data Overview", "Give me an overview of the data structure"),
            ("🎯 Find Patterns", "What patterns can you find in this data?"),
            ("💡 Search Data", "Help me search for specific content in the data"),
            ("📋 Column Analysis", "Analyze each column and its characteristics")
        ]
        
        for i, (label, question) in enumerate(quick_questions):
            col = col1 if i % 2 == 0 else col2
            with col:
                if st.button(label, use_container_width=True):
                    print(f"🚀 Quick question triggered: '{question}'")
                    st.session_state.chat_history.append({"role": "user", "content": question})
                    st.rerun()

if __name__ == "__main__":
    print("🔄 Streamlit app starting...")
    main()